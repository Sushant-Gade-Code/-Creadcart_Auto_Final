import openpyxl


class Xlutil:

     @staticmethod
     def getRowCount(File,SheetName):
         workbook=openpyxl.load_workbook(File)
         sheet=workbook[SheetName]
         return sheet.max_row

     @staticmethod
     def getColumncount(File, SheetName):
         workbook = openpyxl.load_workbook(File)
         sheet = workbook[SheetName]
         return sheet.max_column

     @staticmethod
     def getReadData(File, SheetName, row, column):
         workbook = openpyxl.load_workbook(File)
         sheet = workbook[SheetName]
         return sheet.cell(row, column).value


     @staticmethod
     def getWrightData(File, SheetName,row,column,dada):
         workbook = openpyxl.load_workbook(File)
         sheet = workbook[SheetName]
         (sheet.cell(row,column).value)=dada
         workbook.save(File)







